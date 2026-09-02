"""
PalmaData · Sanidad · Censo de enfermedades · Endpoints
=======================================================
Rutas bajo /api/sanidad. Todas exigen sesión.

El módulo no calcula: la base ya trae las vistas y las funciones de
corrección. Aquí solo se pasan los filtros y se devuelven los datos.
"""
from datetime import date
from io import BytesIO

from fastapi import APIRouter, Body, Depends, HTTPException, Query, Request
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ...core import security
from . import repository as repo

router = APIRouter(prefix="/api/sanidad", tags=["sanidad"])

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
VERDE = "16412B"


def sesion(request: Request) -> dict:
    usuario = security.usuario_actual(request)
    if not usuario:
        raise HTTPException(401, "Sesión no iniciada.")
    return usuario


def _quien(usuario: dict) -> str:
    """
    El username que queda registrado en la corrección.

    Va en texto a propósito: la columna `usuario` de la tabla guarda
    un tipo (1 directivo, 2 trabajador) y los id del login van de 1 a
    15, así que un número ahí sería ambiguo para siempre. El username
    se lee sin cruzar con ninguna otra tabla.
    """
    return usuario["usuario"]


def _filtros(fecha_desde=None, fecha_hasta=None, actualiza_desde=None,
             actualiza_hasta=None, cat_lote_id=None, evaluador=None) -> dict:
    return {"fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta,
            "actualiza_desde": actualiza_desde,
            "actualiza_hasta": actualiza_hasta,
            "cat_lote_id": cat_lote_id, "evaluador": evaluador}


def _exigir_fecha(f: dict):
    """
    Sin filtro de fecha la consulta recorre toda la tabla, que crece
    día a día. Se exige al menos uno de los dos rangos.
    """
    if not any([f.get("fecha_desde"), f.get("fecha_hasta"),
                f.get("actualiza_desde"), f.get("actualiza_hasta")]):
        raise HTTPException(400, "Selecciona un rango de fechas: por fecha "
                                 "del censo o por fecha de actualización.")


def _limpiar(v):
    if isinstance(v, date):
        return v.isoformat()
    if hasattr(v, "quantize"):
        return float(v)
    if hasattr(v, "strftime"):
        return v.strftime("%H:%M:%S")
    return v


def _fila(f: dict) -> dict:
    return {k: _limpiar(v) for k, v in f.items()}


# ============================================================
#  CATÁLOGOS
# ============================================================

@router.get("/catalogos")
def get_catalogos(_=Depends(sesion)):
    """Todo lo que necesitan los desplegables de la ventana de edición."""
    return {"ok": True,
            "enfermedades": repo.listar_enfermedades(),
            "eventos": repo.listar_eventos(),
            "evaluadores": [_fila(x) for x in repo.listar_evaluadores()],
            "fechas": [_fila(x) for x in repo.fechas_disponibles()],
            "actualizaciones": [_fila(x) for x in repo.fechas_actualizacion()]}


@router.get("/lotes")
def get_lotes(q: str | None = Query(None, description="Nombre o número"),
              limite: int = Query(500, ge=1, le=2000), _=Depends(sesion)):
    """
    Lotes para el desplegable. Son unos 500, por eso el buscador:
    escribir «138» encuentra «L138-C» sin recorrer la lista.
    """
    return {"ok": True, "lotes": repo.listar_lotes(q, limite)}


# ============================================================
#  REVISIÓN
# ============================================================

@router.get("/revision")
def get_revision(fecha_desde: date | None = Query(None),
                 fecha_hasta: date | None = Query(None),
                 actualiza_desde: date | None = Query(None),
                 actualiza_hasta: date | None = Query(None),
                 cat_lote_id: int | None = Query(None),
                 evaluador: int | None = Query(None),
                 ver_anulados: bool = Query(False),
                 solo_erroneos: bool = Query(False),
                 limite: int = Query(1000, ge=1, le=10000),
                 _=Depends(sesion)):
    """
    Los registros del censo para revisar y corregir.

    Se puede filtrar por fecha del evento o por fecha de actualización
    —el día en que se descargaron del celular— y además por lote y por
    evaluador.
    """
    f = _filtros(fecha_desde, fecha_hasta, actualiza_desde,
                 actualiza_hasta, cat_lote_id, evaluador)
    _exigir_fecha(f)

    filas = repo.listar_revision(f, ver_anulados, solo_erroneos, limite)
    return {"ok": True, "filtros": {k: _limpiar(v) for k, v in f.items()},
            "total": len(filas), "limite": limite,
            "truncado": len(filas) >= limite,
            "resumen": _fila(repo.resumen(f)),
            "registros": [_fila(x) for x in filas]}


@router.get("/distribucion")
def get_distribucion(campo: str = Query(..., pattern="^(enfermedad|evento|trabajador|lote|fecha)$"),
                     fecha_desde: date | None = Query(None),
                     fecha_hasta: date | None = Query(None),
                     actualiza_desde: date | None = Query(None),
                     actualiza_hasta: date | None = Query(None),
                     cat_lote_id: int | None = Query(None),
                     evaluador: int | None = Query(None),
                     limite: int = Query(20, ge=3, le=100),
                     _=Depends(sesion)):
    f = _filtros(fecha_desde, fecha_hasta, actualiza_desde,
                 actualiza_hasta, cat_lote_id, evaluador)
    _exigir_fecha(f)
    return {"ok": True, "campo": campo,
            "datos": [_fila(x) for x in repo.distribucion(campo, f, limite)]}


@router.get("/duplicados")
def get_duplicados(fecha_desde: date | None = Query(None),
                   fecha_hasta: date | None = Query(None),
                   actualiza_desde: date | None = Query(None),
                   actualiza_hasta: date | None = Query(None),
                   cat_lote_id: int | None = Query(None),
                   evaluador: int | None = Query(None),
                   limite: int = Query(1000, ge=1, le=10000),
                   _=Depends(sesion)):
    """
    Mismo lote, misma línea, misma palma, mismo día: casi siempre un error.

    Ser duplicado depende de OTRAS filas, por eso es una vista y no una
    columna: si de tres repetidos se anulan dos, el que queda deja de
    serlo solo. Una columna seguiría diciendo «duplicado».
    """
    f = _filtros(fecha_desde, fecha_hasta, actualiza_desde,
                 actualiza_hasta, cat_lote_id, evaluador)
    _exigir_fecha(f)
    filas = repo.duplicados(f, limite)

    # Cuántos grupos distintos hay: tres filas repetidas son un caso, no tres.
    # El lote entra en la clave, igual que en la vista.
    grupos = len({(str(x["fecha"]), x["lote"], x["linea"], x["palma"])
                  for x in filas})

    return {"ok": True, "total": len(filas), "grupos": grupos,
            "duplicados": [_fila(x) for x in filas]}


# ============================================================
#  CORRECCIONES
# ============================================================

def _ids(datos: dict) -> list[int]:
    ids = datos.get("ids") or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, "No se seleccionó ningún registro.")
    try:
        return [int(x) for x in ids]
    except (TypeError, ValueError):
        raise HTTPException(400, "Hay identificadores inválidos.")


@router.post("/corregir-lote")
def post_corregir_lote(datos: dict = Body(...), usuario=Depends(sesion)):
    """
    Cambia el lote de uno o varios registros a la vez.

    Es la corrección más frecuente: el evaluador anotó un lote y
    resultó ser otro, y suele pasar con toda una tanda de lecturas.
    """
    ids = _ids(datos)
    cat_lote_id = datos.get("cat_lote_id")
    if not cat_lote_id:
        raise HTTPException(400, "Selecciona el lote correcto.")

    try:
        n = repo.corregir_lote(ids, int(cat_lote_id), _quien(usuario))
    except Exception as e:
        raise HTTPException(400, str(e).split("\n")[0])
    return {"ok": True, "corregidos": n}


@router.post("/corregir")
def post_corregir(datos: dict = Body(...), usuario=Depends(sesion)):
    """
    Corrige un registro campo a campo.

    Los campos que no se envían quedan como estaban: se puede cambiar
    solo la línea sin tocar el resto.
    """
    id_registro = datos.get("id")
    if not id_registro:
        raise HTTPException(400, "Falta el registro a corregir.")

    campos = {k: datos.get(k) for k in
              ("cat_lote_id", "linea", "palma", "san_enfermedades_id",
               "san_evento_enf_id", "observaciones")}
    if all(v in (None, "") for v in campos.values()):
        raise HTTPException(400, "No se envió ningún cambio.")

    campos = {k: (None if v in (None, "") else v) for k, v in campos.items()}

    try:
        n = repo.corregir_registro(int(id_registro), _quien(usuario), campos)
    except Exception as e:
        raise HTTPException(400, str(e).split("\n")[0])
    return {"ok": True, "corregidos": n}


@router.post("/anular")
def post_anular(datos: dict = Body(...), usuario=Depends(sesion)):
    """Marca registros como anulados. No los borra: quedan auditables."""
    ids = _ids(datos)
    try:
        n = repo.anular(ids, _quien(usuario), datos.get("motivo"))
    except Exception as e:
        raise HTTPException(400, str(e).split("\n")[0])
    return {"ok": True, "anulados": n}


@router.post("/reactivar")
def post_reactivar(datos: dict = Body(...), usuario=Depends(sesion)):
    ids = _ids(datos)
    try:
        n = repo.reactivar(ids, _quien(usuario))
    except Exception as e:
        raise HTTPException(400, str(e).split("\n")[0])
    return {"ok": True, "reactivados": n}


# ============================================================
#  DESCARGAS
# ============================================================

COLUMNAS_CONSOLIDADO = [
    ("fecha", "Fecha"), ("hora", "Time"), ("evaluador", "EVALUADOR"),
    ("lote", "LOTE"), ("bloque", "BLOQUE"), ("linea", "LINEA"),
    ("palma", "PALMA"), ("evento", "EVENTO"), ("trabajador", "Trabajador"),
    ("romano", "Romano"), ("observaciones", "Observaciones"),
    ("geom", "GEOM"),   # geometría en texto WKT, solo para el Excel
]


def _excel(titulo: str, columnas: list[tuple], filas: list[dict],
           nota: str | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    ws.append([etiqueta for _clave, etiqueta in columnas])
    for f in filas:
        ws.append([_limpiar(f.get(clave)) for clave, _e in columnas])

    relleno = PatternFill("solid", fgColor=VERDE)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF", size=10)
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    for i, (_c, etiqueta) in enumerate(columnas):
        letra = chr(65 + i) if i < 26 else "A" + chr(65 + i - 26)
        ws.column_dimensions[letra].width = max(12, min(34, len(etiqueta) + 8))
    ws.freeze_panes = "A2"

    if nota:
        guia = wb.create_sheet("filtros")
        for linea in nota.split("\n"):
            guia.append([linea])
        guia.column_dimensions["A"].width = 68
        guia.cell(row=1, column=1).font = Font(bold=True, size=12, color=VERDE)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _nombre(base: str, desde, hasta) -> str:
    partes = [base]
    if desde and hasta and desde == hasta:
        partes.append(str(desde).replace("-", ""))
    else:
        if desde:
            partes.append(str(desde).replace("-", ""))
        if hasta:
            partes.append("a_" + str(hasta).replace("-", ""))
    return "_".join(partes) + ".xlsx"


@router.get("/consolidado")
def get_consolidado(fecha_desde: date | None = Query(None),
                    fecha_hasta: date | None = Query(None),
                    _=Depends(sesion)):
    """Vista previa del consolidado antes de descargarlo."""
    if not fecha_desde and not fecha_hasta:
        raise HTTPException(400, "Selecciona la fecha del censo a consolidar.")
    filas = repo.consolidado(fecha_desde, fecha_hasta)
    return {"ok": True, "total": len(filas),
            "columnas": [e for _c, e in COLUMNAS_CONSOLIDADO],
            "registros": [_fila(x) for x in filas[:500]]}


@router.get("/consolidado/excel")
def get_consolidado_excel(fecha_desde: date | None = Query(None),
                          fecha_hasta: date | None = Query(None),
                          _=Depends(sesion)):
    """
    Descarga el consolidado del censo.

    Va por fecha del EVENTO, no por fecha de descarga: es el censo de
    esos días con las correcciones ya aplicadas.
    """
    if not fecha_desde and not fecha_hasta:
        raise HTTPException(400, "Selecciona la fecha del censo a consolidar.")

    filas = repo.consolidado(fecha_desde, fecha_hasta)
    if not filas:
        raise HTTPException(404, "No hay registros para esas fechas.")

    nota = ("Consolidado del censo de enfermedades\n"
            f"Fecha del censo: {fecha_desde or 'sin límite'} "
            f"a {fecha_hasta or 'sin límite'}\n"
            f"Registros: {len(filas)}\n"
            "No incluye los anulados.")

    contenido = _excel("consolidado", COLUMNAS_CONSOLIDADO, filas, nota)
    archivo = _nombre("censo_consolidado", fecha_desde, fecha_hasta)
    return Response(content=contenido, media_type=XLSX,
                    headers={"Content-Disposition":
                             f'attachment; filename="{archivo}"'})


@router.get("/duplicados/excel")
def get_duplicados_excel(fecha_desde: date | None = Query(None),
                         fecha_hasta: date | None = Query(None),
                         actualiza_desde: date | None = Query(None),
                         actualiza_hasta: date | None = Query(None),
                         cat_lote_id: int | None = Query(None),
                         evaluador: int | None = Query(None),
                         _=Depends(sesion)):
    """Los duplicados del período, para revisarlos fuera de la pantalla."""
    f = _filtros(fecha_desde, fecha_hasta, actualiza_desde,
                 actualiza_hasta, cat_lote_id, evaluador)
    _exigir_fecha(f)

    filas = repo.duplicados(f, 10000)
    columnas = [("fecha", "Fecha"), ("hora", "Hora"), ("lote", "Lote"),
                ("linea", "Linea"), ("palma", "Palma"),
                ("repeticiones", "Veces"),
                ("enfermedad", "Enfermedad"), ("evento", "Evento"),
                ("trabajador", "Trabajador"),
                ("observaciones", "Observaciones"),
                ("id_unico", "ID_unico")]

    nota = ("Registros duplicados del censo\n"
            "Misma palma, misma línea, mismo día.\n"
            f"Fecha del censo: {fecha_desde or '—'} a {fecha_hasta or '—'}\n"
            f"Fecha de actualización: {actualiza_desde or '—'} "
            f"a {actualiza_hasta or '—'}\n"
            f"Registros: {len(filas)}")

    contenido = _excel("duplicados", columnas, filas, nota)
    archivo = _nombre("censo_duplicados",
                      fecha_desde or actualiza_desde,
                      fecha_hasta or actualiza_hasta)
    return Response(content=contenido, media_type=XLSX,
                    headers={"Content-Disposition":
                             f'attachment; filename="{archivo}"'})


@router.get("/revision/excel")
def get_revision_excel(fecha_desde: date | None = Query(None),
                       fecha_hasta: date | None = Query(None),
                       actualiza_desde: date | None = Query(None),
                       actualiza_hasta: date | None = Query(None),
                       cat_lote_id: int | None = Query(None),
                       evaluador: int | None = Query(None),
                       ver_anulados: bool = Query(False),
                       solo_erroneos: bool = Query(False),
                       _=Depends(sesion)):
    """La tabla de revisión tal como se ve en pantalla."""
    f = _filtros(fecha_desde, fecha_hasta, actualiza_desde,
                 actualiza_hasta, cat_lote_id, evaluador)
    _exigir_fecha(f)

    filas = repo.listar_revision(f, ver_anulados, solo_erroneos, 10000)
    columnas = [("fecha", "Fecha"), ("hora", "Hora"), ("lote", "Lote"),
                ("linea", "Linea"), ("palma", "Palma"),
                ("enfermedad", "Enfermedad"), ("evento", "Evento"),
                ("trabajador", "Trabajador"),
                ("observaciones", "Observaciones"),
                ("fecha_actualizacion", "Fecha actualización"),
                ("corregido_por", "Corregido por"),
                ("corregido_at", "Corregido el"),
                ("anulado_por", "Anulado por"),
                ("id_unico", "ID_unico")]

    nota = ("Revisión del censo de enfermedades\n"
            f"Fecha del censo: {fecha_desde or '—'} a {fecha_hasta or '—'}\n"
            f"Fecha de actualización: {actualiza_desde or '—'} "
            f"a {actualiza_hasta or '—'}\n"
            f"Registros: {len(filas)}")

    contenido = _excel("revision", columnas, filas, nota)
    archivo = _nombre("censo_revision",
                      fecha_desde or actualiza_desde,
                      fecha_hasta or actualiza_hasta)
    return Response(content=contenido, media_type=XLSX,
                    headers={"Content-Disposition":
                             f'attachment; filename="{archivo}"'})


# ============================================================
#  TRATAMIENTOS · sub-router del mismo módulo
#  (al final para que main.py siga cargando un solo router)
# ============================================================
from .router_trat import router_trat  # noqa: E402
router.include_router(router_trat)

from .router_plagas import router_plagas  # noqa: E402
router.include_router(router_plagas)

from .router_trampas import router_trampas  # noqa: E402
router.include_router(router_trampas)

from .router_strategus import router_strategus  # noqa: E402
router.include_router(router_strategus)
