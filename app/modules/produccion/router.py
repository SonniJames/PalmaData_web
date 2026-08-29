"""
PalmaData · Producción · Polinización · Endpoints
=================================================
Rutas bajo /api/produccion/poli. Todas exigen sesión.

El módulo no calcula: la base trae las vistas y las funciones. Aquí solo
se pasan filtros y se devuelven datos.

Los ayudantes (sesión, filtros, Excel) están copiados del módulo de
sanidad a propósito: cada módulo es aislado y debe poder arrancar sin
los demás.
"""
from datetime import date
from io import BytesIO

from fastapi import APIRouter, Body, Depends, HTTPException, Query, Request
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ...core import security
from . import repository as repo

router = APIRouter(prefix="/api/produccion", tags=["produccion"])

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
VERDE = "16412B"


def sesion(request: Request) -> dict:
    usuario = security.usuario_actual(request)
    if not usuario:
        raise HTTPException(401, "Sesión no iniciada.")
    return usuario


def _quien(usuario: dict) -> str:
    """
    El username que queda registrado en la corrección.

    Va en texto a propósito: la columna `usuario` de la tabla guarda un
    tipo (1 directivo, 2 trabajador) y los id del login van de 1 a 15,
    así que un número ahí sería ambiguo para siempre.
    """
    return usuario["usuario"]


def _filtros(fecha_desde=None, fecha_hasta=None, actualiza_desde=None,
             actualiza_hasta=None, evaluador=None) -> dict:
    # Sin cat_lote_id: en polinización no hay filtro por lote — los lotes
    # se ven agrupados por trabajador en el informe.
    return {"fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta,
            "actualiza_desde": actualiza_desde,
            "actualiza_hasta": actualiza_hasta,
            "evaluador": evaluador}


def _exigir_fecha(f: dict):
    """Sin filtro de fecha la consulta recorre toda la tabla."""
    if not any([f.get("fecha_desde"), f.get("fecha_hasta"),
                f.get("actualiza_desde"), f.get("actualiza_hasta")]):
        raise HTTPException(400, "Selecciona un rango de fechas: por fecha "
                                 "del evento o por fecha de actualización.")


def _limpiar(v):
    if isinstance(v, date):
        return v.isoformat()
    if hasattr(v, "quantize"):
        return float(v)
    if hasattr(v, "strftime"):
        return v.strftime("%H:%M:%S")
    return v


def _fila(f: dict) -> dict:
    return {k: _limpiar(v) for k, v in f.items()}


def _ids(datos: dict) -> list[int]:
    ids = datos.get("ids") or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, "No se seleccionó ningún registro.")
    try:
        return [int(x) for x in ids]
    except (TypeError, ValueError):
        raise HTTPException(400, "Hay identificadores inválidos.")


def _excel(titulo: str, columnas: list[tuple], filas: list[dict],
           nota: str | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    ws.append([etiqueta for _clave, etiqueta in columnas])
    for f in filas:
        ws.append([_limpiar(f.get(clave)) for clave, _e in columnas])

    relleno = PatternFill("solid", fgColor=VERDE)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF", size=10)
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    for i, (_c, etiqueta) in enumerate(columnas):
        letra = chr(65 + i) if i < 26 else "A" + chr(65 + i - 26)
        ws.column_dimensions[letra].width = max(12, min(34, len(etiqueta) + 8))
    ws.freeze_panes = "A2"

    if nota:
        guia = wb.create_sheet("filtros")
        for linea in nota.split("\n"):
            guia.append([linea])
        guia.column_dimensions["A"].width = 68
        guia.cell(row=1, column=1).font = Font(bold=True, size=12, color=VERDE)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _nombre(base: str, desde, hasta) -> str:
    partes = [base]
    if desde and hasta and desde == hasta:
        partes.append(str(desde).replace("-", ""))
    else:
        if desde:
            partes.append(str(desde).replace("-", ""))
        if hasta:
            partes.append("a_" + str(hasta).replace("-", ""))
    return "_".join(partes) + ".xlsx"


# ============================================================
#  CATÁLOGOS
# ============================================================

@router.get("/poli/catalogos")
def get_catalogos(_=Depends(sesion)):
    """Polinizadores y fechas disponibles para los filtros."""
    return {"ok": True,
            "evaluadores": [_fila(x) for x in repo.listar_evaluadores()],
            "fechas": [_fila(x) for x in repo.fechas_disponibles()],
            "actualizaciones": [_fila(x) for x in repo.fechas_actualizacion()]}


@router.get("/poli/lotes")
def get_lotes(q: str | None = Query(None, description="Nombre o número"),
              limite: int = Query(500, ge=1, le=2000), _=Depends(sesion)):
    """Lotes para el buscador del modal de corrección."""
    return {"ok": True, "lotes": repo.listar_lotes(q, limite)}


# ============================================================
#  INFORME · pantalla de revisión (agrupado, solo lectura)
# ============================================================

@router.get("/poli/informe")
def get_informe(fecha_desde: date | None = Query(None),
                fecha_hasta: date | None = Query(None),
                actualiza_desde: date | None = Query(None),
                actualiza_hasta: date | None = Query(None),
                evaluador: int | None = Query(None),
                _=Depends(sesion)):
    """
    El informe de polinización: una fila por polinizador + fecha, con los
    lotes del día concatenados y las aplicaciones sumadas. Solo lectura:
    la corrección vive en la pantalla de descargas.
    """
    f = _filtros(fecha_desde, fecha_hasta, actualiza_desde,
                 actualiza_hasta, evaluador)
    _exigir_fecha(f)
    filas = repo.informe(f)
    return {"ok": True, "total": len(filas),
            "registros": [_fila(x) for x in filas]}


# ============================================================
#  DETALLE · pantalla de descargas (registro a registro)
# ============================================================

@router.get("/poli/detalle")
def get_detalle(fecha_desde: date | None = Query(None),
                fecha_hasta: date | None = Query(None),
                actualiza_desde: date | None = Query(None),
                actualiza_hasta: date | None = Query(None),
                evaluador: int | None = Query(None),
                ver_anulados: bool = Query(False),
                solo_erroneos: bool = Query(False),
                limite: int = Query(1000, ge=1, le=10000),
                _=Depends(sesion)):
    """
    Los registros uno a uno, para revisar y corregir: con erróneos,
    anulados, y el resumen para las tarjetas.
    """
    f = _filtros(fecha_desde, fecha_hasta, actualiza_desde,
                 actualiza_hasta, evaluador)
    _exigir_fecha(f)

    filas = repo.detalle(f, ver_anulados, solo_erroneos, limite)
    return {"ok": True, "filtros": {k: _limpiar(v) for k, v in f.items()},
            "total": len(filas), "limite": limite,
            "truncado": len(filas) >= limite,
            "resumen": _fila(repo.resumen(f)),
            "registros": [_fila(x) for x in filas]}


# ============================================================
#  CORRECCIONES
# ============================================================

@router.post("/poli/corregir-lote")
def post_corregir_lote(datos: dict = Body(...), usuario=Depends(sesion)):
    """Cambia el lote de uno o varios registros a la vez."""
    ids = _ids(datos)
    cat_lote_id = datos.get("cat_lote_id")
    if not cat_lote_id:
        raise HTTPException(400, "Selecciona el lote correcto.")

    try:
        n = repo.corregir_lote(ids, int(cat_lote_id), _quien(usuario))
    except Exception as e:
        raise HTTPException(400, str(e).split("\n")[0])
    return {"ok": True, "corregidos": n}


@router.post("/poli/corregir")
def post_corregir(datos: dict = Body(...), usuario=Depends(sesion)):
    """
    Corrige un registro campo a campo: lote, línea, palma, las tres
    aplicaciones y observaciones. Lo que no se envía queda como estaba.
    """
    id_registro = datos.get("id")
    if not id_registro:
        raise HTTPException(400, "Falta el registro a corregir.")

    campos = {k: datos.get(k) for k in
              ("cat_lote_id", "linea", "palma", "aplicacion1",
               "aplicacion2", "aplicacion3", "observaciones")}
    if all(v in (None, "") for v in campos.values()):
        raise HTTPException(400, "No se envió ningún cambio.")

    campos = {k: (None if v in (None, "") else v) for k, v in campos.items()}

    try:
        n = repo.corregir_registro(int(id_registro), _quien(usuario), campos)
    except Exception as e:
        raise HTTPException(400, str(e).split("\n")[0])
    return {"ok": True, "corregidos": n}


@router.post("/poli/anular")
def post_anular(datos: dict = Body(...), usuario=Depends(sesion)):
    """Marca registros como anulados. No los borra: quedan auditables."""
    ids = _ids(datos)
    try:
        n = repo.anular(ids, _quien(usuario), datos.get("motivo"))
    except Exception as e:
        raise HTTPException(400, str(e).split("\n")[0])
    return {"ok": True, "anulados": n}


@router.post("/poli/reactivar")
def post_reactivar(datos: dict = Body(...), usuario=Depends(sesion)):
    ids = _ids(datos)
    try:
        n = repo.reactivar(ids, _quien(usuario))
    except Exception as e:
        raise HTTPException(400, str(e).split("\n")[0])
    return {"ok": True, "reactivados": n}


# ============================================================
#  DESCARGAS
# ============================================================

COLUMNAS_CONSOLIDADO = [
    ("fecha", "FECHA"), ("hora", "HORA"), ("lote", "LOTE"),
    ("linea", "LINEA"), ("palma", "PALMA"), ("polinizador", "POLINIZADOR"),
    ("aplicacion1", "APLICACION 1"), ("aplicacion2", "APLICACION 2"),
    ("aplicacion3", "APLICACION 3"),
]


def _exigir_fecha_descarga(fd, fh, ad, ah):
    if not any([fd, fh, ad, ah]):
        raise HTTPException(400, "Selecciona un rango de fechas: por fecha "
                                 "del evento o por fecha de actualización.")


@router.get("/poli/consolidado")
def get_consolidado(fecha_desde: date | None = Query(None),
                    fecha_hasta: date | None = Query(None),
                    actualiza_desde: date | None = Query(None),
                    actualiza_hasta: date | None = Query(None),
                    _=Depends(sesion)):
    """Vista previa del consolidado antes de descargarlo."""
    _exigir_fecha_descarga(fecha_desde, fecha_hasta,
                           actualiza_desde, actualiza_hasta)
    filas = repo.consolidado(fecha_desde, fecha_hasta,
                             actualiza_desde, actualiza_hasta)
    return {"ok": True, "total": len(filas),
            "columnas": [e for _c, e in COLUMNAS_CONSOLIDADO],
            "registros": [_fila(x) for x in filas[:500]]}


@router.get("/poli/consolidado/excel")
def get_consolidado_excel(fecha_desde: date | None = Query(None),
                          fecha_hasta: date | None = Query(None),
                          actualiza_desde: date | None = Query(None),
                          actualiza_hasta: date | None = Query(None),
                          _=Depends(sesion)):
    """
    Descarga el consolidado de polinización, con las correcciones ya
    aplicadas. Filtra por fecha del EVENTO o por fecha de ACTUALIZACIÓN
    (el día en que los registros se descargaron del celular). No incluye
    anulados; los erróneos sí van, marcados en la pantalla para
    corregirlos antes.
    """
    _exigir_fecha_descarga(fecha_desde, fecha_hasta,
                           actualiza_desde, actualiza_hasta)

    filas = repo.consolidado(fecha_desde, fecha_hasta,
                             actualiza_desde, actualiza_hasta)
    if not filas:
        raise HTTPException(404, "No hay registros para esas fechas.")

    nota = ("Consolidado de polinización\n"
            f"Fecha del evento: {fecha_desde or 'sin límite'} "
            f"a {fecha_hasta or 'sin límite'}\n"
            f"Fecha de actualización: {actualiza_desde or 'sin límite'} "
            f"a {actualiza_hasta or 'sin límite'}\n"
            f"Registros: {len(filas)}\n"
            "No incluye los anulados.")

    contenido = _excel("consolidado", COLUMNAS_CONSOLIDADO, filas, nota)
    archivo = _nombre("polinizacion_consolidado",
                      fecha_desde or actualiza_desde,
                      fecha_hasta or actualiza_hasta)
    return Response(content=contenido, media_type=XLSX,
                    headers={"Content-Disposition":
                             f'attachment; filename="{archivo}"'})


@router.get("/poli/detalle/excel")
def get_detalle_excel(fecha_desde: date | None = Query(None),
                      fecha_hasta: date | None = Query(None),
                      actualiza_desde: date | None = Query(None),
                      actualiza_hasta: date | None = Query(None),
                      evaluador: int | None = Query(None),
                      ver_anulados: bool = Query(False),
                      solo_erroneos: bool = Query(False),
                      _=Depends(sesion)):
    """La tabla de trabajo tal como se ve en pantalla."""
    f = _filtros(fecha_desde, fecha_hasta, actualiza_desde,
                 actualiza_hasta, evaluador)
    _exigir_fecha(f)

    filas = repo.detalle(f, ver_anulados, solo_erroneos, 10000)
    columnas = [("fecha", "Fecha"), ("hora", "Hora"), ("lote", "Lote"),
                ("linea", "Linea"), ("palma", "Palma"),
                ("trabajador", "Polinizador"),
                ("aplicacion1", "Aplicacion 1"),
                ("aplicacion2", "Aplicacion 2"),
                ("aplicacion3", "Aplicacion 3"),
                ("inflorescencias", "Total inflorescencias"),
                ("observaciones", "Observaciones"),
                ("erroneo", "Palma inexistente"),
                ("fecha_actualizacion", "Fecha actualización"),
                ("corregido_por", "Corregido por"),
                ("corregido_at", "Corregido el"),
                ("anulado_por", "Anulado por"),
                ("id_unico", "ID_unico")]

    nota = ("Detalle de polinización\n"
            f"Fecha del evento: {fecha_desde or '—'} a {fecha_hasta or '—'}\n"
            f"Fecha de actualización: {actualiza_desde or '—'} "
            f"a {actualiza_hasta or '—'}\n"
            + ("Solo palmas inexistentes\n" if solo_erroneos else "")
            + ("Solo anulados\n" if ver_anulados else "")
            + f"Registros: {len(filas)}")

    contenido = _excel("detalle", columnas, filas, nota)
    archivo = _nombre("polinizacion_detalle",
                      fecha_desde or actualiza_desde,
                      fecha_hasta or actualiza_hasta)
    return Response(content=contenido, media_type=XLSX,
                    headers={"Content-Disposition":
                             f'attachment; filename="{archivo}"'})
