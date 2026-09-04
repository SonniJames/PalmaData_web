"""
PalmaData · Supervisión · Cosecha lote · Endpoints
==================================================
Rutas bajo /api/supervision. Todas exigen sesión.

Apartado de solo consulta: la tabla que se ve es la misma que se descarga.
Sin corrección, sin anulación, sin análisis de erróneos.
"""
from datetime import date, datetime
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ...core import security
from . import repository as repo

router = APIRouter(prefix="/api/supervision", tags=["supervision"])

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
VERDE = "16412B"

# La tabla de la pantalla y la del Excel son la misma, y salen de aquí.
COLUMNAS = [
    ("fecha", "FECHA"), ("hora", "HORA"),
    ("supervisor", "SUPERVISOR"), ("cortador", "CORTADOR"),
    ("recolector", "RECOLECTOR"), ("alistador", "ALISTADOR"),
    ("linea", "LINEA"), ("palma", "PALMA"), ("lote", "LOTE"),
    ("ciclo", "CICLO"),
    ("racimos_sin_recoger", "RACIMOS SIN RECOGER"),
    ("racimos_sin_cortar", "RACIMOS SIN CORTAR"),
    ("racimo_robado", "RACIMO ROBADO"),
    ("hojas_mal_acomodadas", "HOJAS MAL ACOMODADAS"),
    ("hoja_colgando", "HOJA COLGANDO"),
    ("fruto_plato", "FRUTO PLATO"),
    ("observaciones", "OBSERVACIONES"),
    ("racimos_recogidos", "RACIMOS RECOGIDOS"),
    ("racimos_verdes", "RACIMOS VERDES"),
    ("racimos_sobremaduros", "RACIMOS SOBREMADUROS"),
    ("racimos_podridos", "RACIMOS PODRIDOS"),
]

ROLES = ("supervisor", "cortador", "recolector", "alistador")


def sesion(request: Request) -> dict:
    usuario = security.usuario_actual(request)
    if not usuario:
        raise HTTPException(401, "Sesión no iniciada.")
    return usuario


def _limpiar(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M")
    if isinstance(v, date):
        return v.isoformat()
    if hasattr(v, "quantize"):
        return float(v)
    if hasattr(v, "strftime"):
        return v.strftime("%H:%M")
    return v


def _fila(f: dict) -> dict:
    return {k: _limpiar(v) for k, v in f.items()}


def _filtros(fecha_desde=None, fecha_hasta=None, actualiza_desde=None,
             actualiza_hasta=None, cat_lote_id=None, supervisor=None,
             cortador=None, recolector=None, alistador=None) -> dict:
    return {"fecha_desde": fecha_desde, "fecha_hasta": fecha_hasta,
            "actualiza_desde": actualiza_desde, "actualiza_hasta": actualiza_hasta,
            "cat_lote_id": cat_lote_id, "supervisor": supervisor,
            "cortador": cortador, "recolector": recolector,
            "alistador": alistador}


def _exigir_fecha(f: dict):
    """Sin fecha la consulta recorrería toda la tabla."""
    if not any([f.get("fecha_desde"), f.get("fecha_hasta"),
                f.get("actualiza_desde"), f.get("actualiza_hasta")]):
        raise HTTPException(400, "Selecciona un rango de fechas: por fecha "
                                 "del evento o por fecha de actualización.")


def _excel(titulo: str, columnas: list[tuple], filas: list[dict],
           nota: str | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    ws.append([etiqueta for _clave, etiqueta in columnas])
    for f in filas:
        ws.append([_limpiar(f.get(clave)) for clave, _e in columnas])

    relleno = PatternFill("solid", fgColor=VERDE)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF", size=10)
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    for i, (_c, etiqueta) in enumerate(columnas):
        letra = chr(65 + i) if i < 26 else "A" + chr(65 + i - 26)
        ws.column_dimensions[letra].width = max(12, min(34, len(etiqueta) + 6))
    ws.freeze_panes = "A2"

    if nota:
        guia = wb.create_sheet("filtros")
        for linea in nota.split("\n"):
            guia.append([linea])
        guia.column_dimensions["A"].width = 68
        guia.cell(row=1, column=1).font = Font(bold=True, size=12, color=VERDE)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _nombre(base: str, desde, hasta) -> str:
    partes = [base]
    if desde and hasta and desde == hasta:
        partes.append(str(desde).replace("-", ""))
    else:
        if desde:
            partes.append(str(desde).replace("-", ""))
        if hasta:
            partes.append("a_" + str(hasta).replace("-", ""))
    return "_".join(partes) + ".xlsx"


# ============================================================
#  CATÁLOGOS
# ============================================================

@router.get("/cosecha/catalogos")
def get_catalogos(_=Depends(sesion)):
    """Trabajadores por rol, lotes y fechas disponibles."""
    return {"ok": True,
            "personas": {rol: [_fila(x) for x in repo.personas(rol)] for rol in ROLES},
            "lotes": [_fila(x) for x in repo.lotes()],
            "fechas": [_fila(x) for x in repo.fechas_disponibles()],
            "actualizaciones": [_fila(x) for x in repo.fechas_actualizacion()]}


# ============================================================
#  CONSULTA Y DESCARGA
# ============================================================

@router.get("/cosecha")
def get_cosecha(fecha_desde: date | None = Query(None),
                fecha_hasta: date | None = Query(None),
                actualiza_desde: date | None = Query(None),
                actualiza_hasta: date | None = Query(None),
                cat_lote_id: int | None = Query(None),
                supervisor: int | None = Query(None),
                cortador: int | None = Query(None),
                recolector: int | None = Query(None),
                alistador: int | None = Query(None),
                limite: int = Query(5000, ge=1, le=50000),
                _=Depends(sesion)):
    """
    Los registros de supervisión de cosecha. Los filtros de cortador,
    recolector y alistador traen los registros donde esa persona aparezca,
    sola o acompañada.
    """
    f = _filtros(fecha_desde, fecha_hasta, actualiza_desde, actualiza_hasta,
                 cat_lote_id, supervisor, cortador, recolector, alistador)
    _exigir_fecha(f)

    filas = repo.listar(f, limite)
    return {"ok": True, "filtros": {k: _limpiar(v) for k, v in f.items()},
            "total": len(filas), "limite": limite,
            "truncado": len(filas) >= limite,
            "columnas": [e for _c, e in COLUMNAS],
            "resumen": _fila(repo.resumen(f)),
            "registros": [_fila(x) for x in filas]}


@router.get("/cosecha/excel")
def get_cosecha_excel(fecha_desde: date | None = Query(None),
                      fecha_hasta: date | None = Query(None),
                      actualiza_desde: date | None = Query(None),
                      actualiza_hasta: date | None = Query(None),
                      cat_lote_id: int | None = Query(None),
                      supervisor: int | None = Query(None),
                      cortador: int | None = Query(None),
                      recolector: int | None = Query(None),
                      alistador: int | None = Query(None),
                      _=Depends(sesion)):
    """La misma tabla que se ve en pantalla, en Excel."""
    f = _filtros(fecha_desde, fecha_hasta, actualiza_desde, actualiza_hasta,
                 cat_lote_id, supervisor, cortador, recolector, alistador)
    _exigir_fecha(f)

    filas = repo.listar(f, 50000)
    if not filas:
        raise HTTPException(404, "No hay registros para esos filtros.")

    r = repo.resumen(f)
    nota = ("Supervisión de cosecha por lote\n"
            f"Fecha del evento: {fecha_desde or 'sin límite'} a {fecha_hasta or 'sin límite'}\n"
            f"Fecha de actualización: {actualiza_desde or 'sin límite'} "
            f"a {actualiza_hasta or 'sin límite'}\n"
            f"Registros: {len(filas)}\n"
            f"Ciclo promedio: {r.get('ciclo_promedio') or '—'}\n"
            "Los filtros de cortador, recolector y alistador incluyen los "
            "registros donde la persona aparece acompañada.")

    contenido = _excel("cosecha", COLUMNAS, filas, nota)
    archivo = _nombre("supervision_cosecha",
                      fecha_desde or actualiza_desde, fecha_hasta or actualiza_hasta)
    return Response(content=contenido, media_type=XLSX,
                    headers={"Content-Disposition": f'attachment; filename="{archivo}"'})


# ============================================================
#  COSECHA VAGÓN · se monta al final, cuando los ayudantes ya existen
# ============================================================
from .router_vagon import router_vagon  # noqa: E402
router.include_router(router_vagon)


# ============================================================
#  POLINIZACIÓN
# ============================================================
from .router_poli import router_poli  # noqa: E402
router.include_router(router_poli)
