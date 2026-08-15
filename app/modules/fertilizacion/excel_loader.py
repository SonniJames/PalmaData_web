"""
PalmaData · Fertilización · Carga y formato del Excel
=====================================================
Lee el archivo por HOJAS. Cada hoja sigue la misma regla:
fila 1 = nombres de columna, columna A = identificación del lote.

Ninguna hoja tiene columnas fijas: lo que traiga el archivo se guarda.
Así, si una campaña usa otros fertilizantes u otros nutrientes,
el sistema los reconoce sin cambios de código ni de tablas.
"""
import unicodedata
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import formato as F


# ============================================================
#  Utilidades
# ============================================================

def _norm(texto) -> str:
    """minúsculas, sin acentos, sin espacios ni símbolos."""
    if texto is None:
        return ""
    t = unicodedata.normalize("NFKD", str(texto).strip().lower())
    t = "".join(c for c in t if not unicodedata.combining(c))
    for ch in (" ", "_", ".", "-", "/", "(", ")", "°", "%"):
        t = t.replace(ch, "")
    return t


def _txt(valor):
    if valor is None:
        return None
    t = str(valor).strip()
    return t or None


ERRORES_EXCEL = {"#n/a", "#¡div/0!", "#div/0!", "#value!", "#¡valor!",
                 "#ref!", "#¡ref!", "#name?", "#¿nombre?", "-", "—", ""}


def _num(valor, entero=False):
    """Convierte a número. Vacío, texto o error de Excel -> None."""
    if valor is None:
        return None
    if isinstance(valor, str):
        v = valor.strip().replace(",", ".")
        if v.lower() in ERRORES_EXCEL:
            return None
        valor = v
    try:
        f = float(valor)
    except (TypeError, ValueError):
        return None
    if f != f or f in (float("inf"), float("-inf")):
        return None
    return int(round(f)) if entero else f


def _buscar_hoja(wb, clave: str):
    """Encuentra una hoja por su nombre o cualquiera de sus alias."""
    alias = {_norm(a) for a in F.ALIAS_HOJAS.get(clave, [clave])}
    for nombre in wb.sheetnames:
        if _norm(nombre) in alias:
            return wb[nombre]
    return None


def _leer_tabla(ws) -> tuple[dict[str, dict], list[str]]:
    """
    Lee una hoja con el formato estándar.
    Devuelve ({identificacion: {columna: valor}}, columnas_en_orden).
    """
    filas = list(ws.iter_rows(values_only=True))
    if len(filas) < 2:
        return {}, []

    cabecera = filas[0]
    columnas = []
    for i, h in enumerate(cabecera):
        if i == 0:
            continue                       # columna A = identificación
        nombre = _txt(h)
        if nombre:
            columnas.append((i, nombre))

    tabla: dict[str, dict] = {}
    for fila in filas[1:]:
        if fila is None:
            continue
        ident = _txt(fila[0]) if len(fila) > 0 else None
        if not ident:
            continue
        tabla[ident] = {nombre: (fila[i] if i < len(fila) else None)
                        for i, nombre in columnas}

    return tabla, [n for _, n in columnas]


# ============================================================
#  Lectura del archivo
# ============================================================

def leer_excel(contenido: bytes) -> tuple[dict, list[str]]:
    """
    Devuelve (resultado, advertencias).

    resultado = {
      "lotes": [ {identificacion, uma, sector, ..., extra,
                  bloques: {foliar:{}, balance:{}, requerimiento:{}}} ],
      "columnas": {"foliar": [...], "balance": [...], "requerimiento": [...]},
      "hojas_leidas": [...],
    }
    """
    advertencias: list[str] = []

    try:
        wb = load_workbook(BytesIO(contenido), data_only=True)
    except Exception as e:
        return {}, [f"No se pudo abrir el archivo: {e}"]

    # ---- Hoja de identificación (obligatoria) ----
    ws_ident = _buscar_hoja(wb, F.HOJA_IDENTIFICACION)
    if ws_ident is None:
        return {}, [f"Falta la hoja «{F.HOJA_IDENTIFICACION}». "
                    f"Descarga el formato desde el módulo y úsalo como base."]

    filas = list(ws_ident.iter_rows(values_only=True))
    if len(filas) < 2:
        return {}, [f"La hoja «{F.HOJA_IDENTIFICACION}» no tiene datos."]

    cabecera = filas[0]
    mapa: dict[int, str] = {}
    extras: dict[int, str] = {}
    for i, h in enumerate(cabecera):
        nombre = _txt(h)
        if not nombre:
            continue
        campo = F.CAMPOS_IDENTIFICACION.get(_norm(nombre))
        if campo:
            mapa[i] = campo
        else:
            extras[i] = nombre

    if "identificacion" not in mapa.values():
        mapa[0] = "identificacion"       # por convención, la columna A
        extras.pop(0, None)

    lotes: list[dict] = []
    vistos: set[str] = set()

    for n, fila in enumerate(filas[1:], start=2):
        if fila is None:
            continue

        reg: dict = {"fila_excel": n, "extra": {}}
        for i, campo in mapa.items():
            valor = fila[i] if i < len(fila) else None
            if campo in F.TEXTO:
                reg[campo] = _txt(valor)
            elif campo in F.ENTEROS:
                reg[campo] = _num(valor, entero=True)
            else:
                reg[campo] = _num(valor)

        ident = reg.get("identificacion")
        if not ident:
            continue

        clave = ident.lower()
        if clave in vistos:
            advertencias.append(f"Fila {n}: «{ident}» está repetido. Se omite.")
            continue
        vistos.add(clave)

        for i, nombre in extras.items():
            valor = fila[i] if i < len(fila) else None
            if valor is not None:
                reg["extra"][nombre] = _num(valor) if _num(valor) is not None else _txt(valor)

        reg["bloques"] = {}
        lotes.append(reg)

    if not lotes:
        return {}, ["No se encontró ningún lote. Revisa que la columna A "
                    "tenga las identificaciones y la fila 1 los nombres de columna."]

    # ---- Hojas de datos ----
    columnas: dict[str, list[str]] = {}
    hojas_leidas = [F.HOJA_IDENTIFICACION]
    por_ident = {l["identificacion"].lower(): l for l in lotes}

    for clave, (nombre_hoja, _tabla, etiqueta) in F.HOJAS_DATOS.items():
        ws = _buscar_hoja(wb, nombre_hoja)
        if ws is None:
            advertencias.append(
                f"No se encontró la hoja «{nombre_hoja}» ({etiqueta}). "
                f"Esa sección quedará vacía.")
            columnas[clave] = []
            continue

        tabla, cols = _leer_tabla(ws)
        columnas[clave] = cols
        hojas_leidas.append(nombre_hoja)

        if not cols:
            advertencias.append(f"La hoja «{nombre_hoja}» no tiene columnas de datos.")
            continue

        sin_lote = []
        for ident, valores in tabla.items():
            lote = por_ident.get(ident.lower())
            if lote is None:
                sin_lote.append(ident)
                continue
            limpio = {}
            for col, valor in valores.items():
                v = _num(valor)
                if v is not None:
                    limpio[col] = v
            lote["bloques"][clave] = limpio

        if sin_lote:
            advertencias.append(
                f"Hoja «{nombre_hoja}»: {len(sin_lote)} identificación(es) no existen "
                f"en la hoja identificacion y se omitieron: "
                f"{', '.join(sin_lote[:4])}" + ("…" if len(sin_lote) > 4 else ""))

        faltantes = [l["identificacion"] for l in lotes
                     if clave not in l["bloques"]]
        if faltantes:
            advertencias.append(
                f"Hoja «{nombre_hoja}»: {len(faltantes)} lote(s) sin datos: "
                f"{', '.join(faltantes[:4])}" + ("…" if len(faltantes) > 4 else ""))

    return {"lotes": lotes, "columnas": columnas,
            "hojas_leidas": hojas_leidas}, advertencias


# ============================================================
#  Generación del formato en blanco
# ============================================================

VERDE = "16412B"
VERDE2 = "2F7D4F"
CREMA = "EAE5D9"


def _estilo_cabecera(ws, ncols: int, anchos=None):
    relleno = PatternFill("solid", fgColor=VERDE)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF", size=10)
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
    ws.row_dimensions[1].height = 32
    for i in range(ncols):
        letra = get_column_letter(i + 1)
        ws.column_dimensions[letra].width = (anchos or {}).get(i, 30 if i == 0 else 13)
    ws.freeze_panes = "B2"


def generar_formato(identificaciones: list[str] | None = None,
                    fertilizantes: list[str] | None = None,
                    empresa: str | None = None) -> bytes:
    """
    Genera el Excel de formato con sus seis hojas de datos más instrucciones.

    Si se pasan `identificaciones` (por ejemplo, las de la campaña anterior),
    se precargan en la columna A de todas las hojas, así el usuario solo
    pega los valores.
    """
    idents = identificaciones or []
    ferts = fertilizantes or ["Grado 13-5-27-5(Mg)", "NCa", "Rafos",
                              "KSOMgO", "KIESE", "Borax 48%", "ZnSO4"]

    wb = Workbook()

    # ---------- INSTRUCCIONES ----------
    guia = wb.active
    guia.title = F.HOJA_INSTRUCCIONES
    lineas = [
        ("PalmaData · Formato de carga · Módulo Fertilización", "titulo"),
        ("", ""),
        ("Regla general para TODAS las hojas", "sub"),
        ("· Fila 1: los nombres de las columnas.", ""),
        ("· Columna A: la identificación del lote. Es la llave que une las hojas.", ""),
        ("· De la columna B en adelante: los valores.", ""),
        ("· La identificación debe escribirse IGUAL en todas las hojas.", ""),
        ("", ""),
        ("Hoja  identificacion", "sub"),
        ("Quién es cada lote. Columnas reconocidas:", ""),
        ("   identificacion · empresa · uma · sector · zona · rango_edad ·", ""),
        ("   palmas · hectareas · material · siembra · codigo · hoja · mst · tons", ""),
        ("La columna empresa es opcional pero recomendada: la empresa se elige", ""),
        ("al cargar, y si el archivo la trae, el sistema verifica que coincida.", ""),
        ("Así te avisa si estás subiendo el archivo equivocado.", ""),
        ("Cualquier columna adicional se guarda igual y queda disponible.", ""),
        ("La columna hectareas es opcional: si la llenas, el sistema calcula", ""),
        ("el costo por hectárea por zona y por sector.", ""),
        ("", ""),
        ("Hoja  anal_foliar", "sub"),
        ("Resultado del laboratorio. Una columna por nutriente (N, P, K, ...).", ""),
        ("Alimenta la pantalla de Diagnóstico.", ""),
        ("", ""),
        ("Hoja  ind_balan", "sub"),
        ("Índice de balance: porcentaje sobre el nivel óptimo.", ""),
        ("Una columna por nutriente. Alimenta el semáforo y el estado nutricional.", ""),
        ("", ""),
        ("Hoja  reque_fert", "sub"),
        ("Fertilizantes requeridos y su cantidad por lote.", ""),
        ("Una columna por fertilizante, con el nombre del producto en la fila 1.", ""),
        ("Puedes agregar, quitar o cambiar fertilizantes entre campañas:", ""),
        ("el sistema los detecta solos y los precios aparecen en Parámetros.", ""),
        ("", ""),
        ("Hoja  reque_ox", "sub"),
        ("Requerimiento expresado en forma de ÓXIDO, que es como se", ""),
        ("comercializan los fertilizantes: P2O5, K2O, CaO, MgO, B2O3.", ""),
        ("Una columna por elemento. Alimenta la pantalla Requerimiento en óxido.", ""),
        ("", ""),
        ("Hoja  reque_rend", "sub"),
        ("REQUERIMIENTO TOTAL PARA EL RENDIMIENTO ESPERADO.", ""),
        ("Nutrientes en forma elemental (N, P, K, Ca, Mg, S, B, Zn) que hacen", ""),
        ("falta para alcanzar la cosecha esperada de cada lote.", ""),
        ("Alimenta la pantalla Requerimiento para rendimiento.", ""),
        ("", ""),
        ("En ambas, como en todas las hojas, la columna A lleva la", ""),
        ("identificación del lote y los elementos pueden cambiar entre", ""),
        ("campañas: el sistema los detecta solos.", ""),
        ("", ""),
        ("Cómo cargar", "sub"),
        ("1. Copia tus datos y pégalos aquí con Pegado especial → Valores.", ""),
        ("2. No cambies los nombres de las hojas.", ""),
        ("3. Sube el archivo desde Fertilización → Cargar datos, eligiendo", ""),
        ("   la EMPRESA y el año.", ""),
        ("4. Recargar la misma empresa y año reemplaza esos datos: si te", ""),
        ("   equivocaste de archivo, sube el correcto y queda limpio.", ""),
        ("   Los precios y fletes que ya hayas puesto NO se borran.", ""),
        ("", ""),
        ("Qué calcula PalmaData", "sub"),
        ("El sistema NO recalcula la agronomía: guarda tus valores tal cual.", ""),
        ("Calcula los totales por zona, sector y edad, los costos", ""),
        ("(cantidad × precio), el costo por palma y por hectárea, y las gráficas.", ""),
        ("Los precios se ingresan en la pestaña Parámetros, por campaña.", ""),
    ]
    for texto, tipo in lineas:
        guia.append([texto])
        celda = guia.cell(row=guia.max_row, column=1)
        if tipo == "titulo":
            celda.font = Font(bold=True, size=14, color=VERDE)
        elif tipo == "sub":
            celda.font = Font(bold=True, size=11, color=VERDE2)
    guia.column_dimensions["A"].width = 86

    # ---------- identificacion ----------
    ws = wb.create_sheet(F.HOJA_IDENTIFICACION)
    cols_ident = ["identificacion", "empresa", "uma", "sector", "zona",
                  "rango_edad", "palmas", "hectareas", "material", "siembra",
                  "codigo", "hoja", "mst", "tons"]
    ws.append(cols_ident)
    for ident in idents:
        ws.append([ident, empresa])
    _estilo_cabecera(ws, len(cols_ident), {0: 34, 1: 22, 8: 24})

    # ---------- anal_foliar ----------
    ws = wb.create_sheet(F.HOJA_FOLIAR)
    ws.append(["identificacion"] + F.ORDEN_NUTRIENTES)
    for ident in idents:
        ws.append([ident])
    _estilo_cabecera(ws, len(F.ORDEN_NUTRIENTES) + 1, {0: 34})

    # ---------- ind_balan ----------
    ws = wb.create_sheet(F.HOJA_BALANCE)
    nut_bal = [n for n in F.ORDEN_NUTRIENTES if n != "Cl"]
    ws.append(["identificacion"] + nut_bal)
    for ident in idents:
        ws.append([ident])
    _estilo_cabecera(ws, len(nut_bal) + 1, {0: 34})

    # ---------- reque_fert ----------
    ws = wb.create_sheet(F.HOJA_REQUERIMIENTO)
    ws.append(["identificacion"] + ferts)
    for ident in idents:
        ws.append([ident])
    _estilo_cabecera(ws, len(ferts) + 1, {0: 34})
    for i in range(len(ferts)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 20

    # ---------- reque_ox ----------
    ws = wb.create_sheet(F.HOJA_OXIDO)
    oxidos = ["N", "P2O5", "K2O", "CaO", "MgO", "S", "B2O3"]
    ws.append(["identificacion"] + oxidos)
    for ident in idents:
        ws.append([ident])
    _estilo_cabecera(ws, len(oxidos) + 1, {0: 34})

    # ---------- reque_rend ----------
    ws = wb.create_sheet(F.HOJA_RENDIMIENTO)
    elementos = ["N", "P", "K", "Ca", "Mg", "S", "B", "Zn"]
    ws.append(["identificacion"] + elementos)
    for ident in idents:
        ws.append([ident])
    _estilo_cabecera(ws, len(elementos) + 1, {0: 34})

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
