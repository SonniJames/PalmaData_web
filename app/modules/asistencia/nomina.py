"""
PalmaData · Asistencia · Nómina de trabajadores activos
=======================================================
Una sola tabla para todas las empresas. La empresa viene en el
archivo, así que no se elige al cargar.

Columnas del Excel:
    Codigo · Nombre Del Trabajador · Employee ID · estado ·
    id · supervisor · empresa

La columna `id` es la llave del cruce: "EmployeeID_Nombre", con el
nombre tal como aparece en el huellero. El Employee ID solo no basta,
porque se repite entre personas distintas: en Villa Claudia hay dos
trabajadores con el ID 163.

    163_Cristian Moreno  ->  Cristian Danilo Moreno Martinez
    163_Carlos Gomez     ->  Carlos Gomez

El cruce es EXACTO, sin normalizar: el `id` se construye copiando el
nombre del huellero, así que debe coincidir carácter por carácter.
"""
import unicodedata
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

VERDE = "16412B"
VERDE2 = "2F7D4F"

COLUMNAS = ["Codigo", "Nombre Del Trabajador", "Employee ID",
            "estado", "id", "supervisor", "empresa"]


def _norm(texto) -> str:
    """Solo para reconocer ENCABEZADOS, nunca para cruzar datos."""
    if texto is None:
        return ""
    t = unicodedata.normalize("NFKD", str(texto).strip().lower())
    t = "".join(c for c in t if not unicodedata.combining(c))
    for ch in (" ", "_", ".", "-", "/"):
        t = t.replace(ch, "")
    return t


def _txt(v):
    """Texto tal cual, sin normalizar: el cruce depende de esto."""
    if v is None:
        return None
    t = str(v).strip()
    return t or None


def _entero(v):
    if v is None:
        return None
    try:
        return int(float(str(v).strip()))
    except (TypeError, ValueError):
        return None


MAPA = {
    "codigo": "codigo", "cod": "codigo", "codigonomina": "codigo",
    "nombredeltrabajador": "nombre", "nombre": "nombre",
    "trabajador": "nombre", "nombrecompleto": "nombre",
    "employeeid": "employee_id", "idhuellero": "employee_id",
    "estado": "estado",
    "id": "id_compuesto", "idcompuesto": "id_compuesto",
    "supervisor": "supervisor", "jefe": "supervisor",
    "empresa": "empresa_id", "empresaid": "empresa_id",
    "idempresa": "empresa_id",
}


def leer_nomina(contenido: bytes) -> tuple[list[dict], list[str]]:
    """
    Lee el Excel de trabajadores activos.

    Devuelve (registros, advertencias). Los registros sin `id`
    compuesto se cargan igual, pero se avisa: nunca van a cruzar
    con una marcación.
    """
    advertencias: list[str] = []

    try:
        wb = load_workbook(BytesIO(contenido), data_only=True)
    except Exception as e:
        return [], [f"No se pudo abrir el archivo: {e}"]

    # Preferir la hoja de datos si el archivo trae varias
    preferidas = {"trabajadores", "tablacargar", "nomina", "activos", "datos"}
    ws = None
    for nombre_hoja in wb.sheetnames:
        if _norm(nombre_hoja) in preferidas:
            ws = wb[nombre_hoja]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    filas = list(ws.iter_rows(values_only=True))
    if len(filas) < 2:
        return [], ["El archivo no tiene filas de datos."]

    columnas: dict[int, str] = {}
    for i, valor in enumerate(filas[0]):
        campo = MAPA.get(_norm(valor))
        if campo and campo not in columnas.values():
            columnas[i] = campo

    faltan = [c for c in ("codigo", "nombre", "empresa_id")
              if c not in columnas.values()]
    if faltan:
        etiquetas = {"codigo": "Codigo", "nombre": "Nombre Del Trabajador",
                     "empresa_id": "empresa"}
        return [], [f"Faltan columnas obligatorias: "
                    f"{', '.join(etiquetas[c] for c in faltan)}. "
                    f"Descarga el formato desde el módulo y úsalo como base."]

    registros: list[dict] = []
    vistos: set[tuple] = set()
    sin_id = 0
    repetidos = 0

    for n, fila in enumerate(filas[1:], start=2):
        if fila is None or all(v in (None, "") for v in fila):
            continue

        reg = {campo: (fila[i] if i < len(fila) else None)
               for i, campo in columnas.items()}

        codigo = _txt(reg.get("codigo"))
        nombre = _txt(reg.get("nombre"))
        empresa_id = _entero(reg.get("empresa_id"))
        if not codigo or not nombre or not empresa_id:
            continue

        id_compuesto = _txt(reg.get("id_compuesto"))
        if id_compuesto and id_compuesto.lower() == "none":
            id_compuesto = None
        if not id_compuesto:
            sin_id += 1

        # Un id compuesto no puede apuntar a dos personas
        if id_compuesto:
            clave = (empresa_id, id_compuesto)
            if clave in vistos:
                repetidos += 1
                advertencias.append(
                    f"Fila {n}: el id «{id_compuesto}» ya estaba en la empresa "
                    f"{empresa_id}. Se conserva el primero.")
                continue
            vistos.add(clave)

        estado = _entero(reg.get("estado"))
        registros.append({
            "empresa_id": empresa_id,
            "codigo": codigo,
            "nombre": nombre,
            "employee_id": _txt(reg.get("employee_id")),
            "id_compuesto": id_compuesto,
            "supervisor": _txt(reg.get("supervisor")),
            "estado": 1 if estado is None else estado,
            "fila_excel": n,
        })

    if not registros:
        return [], ["No se encontró ningún trabajador válido. Revisa que estén "
                    "las columnas Codigo, Nombre Del Trabajador y empresa."]

    if sin_id:
        advertencias.append(
            f"{sin_id} trabajador(es) sin la columna «id» (EmployeeID_Nombre). "
            f"Se cargan igual, pero no cruzarán con ninguna marcación hasta "
            f"que se les complete.")

    empresas = sorted({r["empresa_id"] for r in registros})
    advertencias.insert(0, f"Empresas en el archivo: "
                           f"{', '.join(str(e) for e in empresas)}.")

    return registros, advertencias


# ============================================================
#  Formato en blanco
# ============================================================

def generar_formato(registros: list[dict] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "trabajadores"
    ws.append(COLUMNAS)

    for r in (registros or []):
        ws.append([r.get("codigo"), r.get("nombre"), r.get("employee_id"),
                   r.get("estado", 1), r.get("id_compuesto"),
                   r.get("supervisor"), r.get("empresa_id")])

    relleno = PatternFill("solid", fgColor=VERDE)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF", size=10)
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    for letra, ancho in (("A", 12), ("B", 38), ("C", 14), ("D", 9),
                         ("E", 30), ("F", 26), ("G", 10)):
        ws.column_dimensions[letra].width = ancho
    ws.freeze_panes = "A2"

    guia = wb.create_sheet("instrucciones")
    lineas = [
        ("PalmaData · Asistencia · Trabajadores activos", "titulo"),
        ("", ""),
        ("Qué es esta carga", "sub"),
        ("La lista de quienes trabajan hoy, de TODAS las empresas en un", ""),
        ("solo archivo. Cada carga reemplaza por completo la anterior.", ""),
        ("", ""),
        ("Sirve para que los análisis del módulo cuenten solo a quienes", ""),
        ("deben marcar, sin la gente que sigue registrada en el huellero", ""),
        ("pero ya se fue.", ""),
        ("", ""),
        ("Columnas", "sub"),
        ("· Codigo: el código de nómina. Obligatorio.", ""),
        ("· Nombre Del Trabajador: nombre completo y bien escrito. Obligatorio.", ""),
        ("· Employee ID: el id con el que quedó registrado en el huellero.", ""),
        ("· estado: 1 para los activos.", ""),
        ("· id: EmployeeID_Nombre, con el nombre TAL COMO viene del huellero.", ""),
        ("· supervisor: el jefe a cargo. Puede ir vacío por ahora.", ""),
        ("· empresa: 1 = Palmeras de Yarima · 2 = Villa Claudia · 3 = CUCÚ", ""),
        ("", ""),
        ("La columna id es la llave del cruce", "sub"),
        ("El Employee ID por sí solo no alcanza: se repite entre personas", ""),
        ("distintas. En Villa Claudia hay dos trabajadores con el ID 163:", ""),
        ("", ""),
        ("   163_Cristian Moreno   ->  Cristian Danilo Moreno Martinez", ""),
        ("   163_Carlos Gomez      ->  Carlos Gomez", ""),
        ("", ""),
        ("El cruce es EXACTO: el nombre debe coincidir carácter por", ""),
        ("carácter con el del huellero, incluidos espacios y mayúsculas.", ""),
        ("", ""),
        ("Cómo funciona el proceso", "sub"),
        ("1. Primero se carga esta tabla. Sin ella no se pueden subir", ""),
        ("   asistencias: no habría contra qué cruzar.", ""),
        ("2. Al cargar un archivo del huellero, cada marcación se cruza", ""),
        ("   contra esta lista y el resultado queda GUARDADO en el registro.", ""),
        ("3. Si dentro de dos meses alguien sale de la nómina, sus", ""),
        ("   marcaciones de hoy siguen contando: reflejan lo que pasaba", ""),
        ("   en ese momento. Solo las nuevas quedarán como inactivas.", ""),
        ("", ""),
        ("Si un trabajador no aparece en los análisis, revisa su columna", ""),
        ("id: probablemente no coincide con el nombre del huellero.", ""),
        ("El módulo lista los que no cruzaron para que puedas corregirlos.", ""),
    ]
    for texto, tipo in lineas:
        guia.append([texto])
        celda = guia.cell(row=guia.max_row, column=1)
        if tipo == "titulo":
            celda.font = Font(bold=True, size=14, color=VERDE)
        elif tipo == "sub":
            celda.font = Font(bold=True, size=11, color=VERDE2)
    guia.column_dimensions["A"].width = 78

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
#  Exportar resultados a Excel
# ============================================================

def exportar_tabla(titulo: str, columnas: list[str],
                   filas: list[list], nota: str | None = None) -> bytes:
    """Genera un Excel simple con una tabla, para descargar del módulo."""
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31] or "datos"

    ws.append(columnas)
    for fila in filas:
        ws.append(fila)

    relleno = PatternFill("solid", fgColor=VERDE)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF", size=10)
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    anchos = [max(12, min(38, len(str(c)) + 6)) for c in columnas]
    for i, ancho in enumerate(anchos):
        ws.column_dimensions[chr(65 + i) if i < 26 else "A"].width = ancho
    ws.freeze_panes = "A2"

    if nota:
        guia = wb.create_sheet("filtros")
        for linea in nota.split("\n"):
            guia.append([linea])
        guia.column_dimensions["A"].width = 70
        guia.cell(row=1, column=1).font = Font(bold=True, size=12, color=VERDE)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
