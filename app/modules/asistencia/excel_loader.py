"""
PalmaData · Asistencia · Carga del Excel del huellero
=====================================================
El archivo viene con una fila por trabajador y una columna por día:

    Employee ID | Name | 1 | 2 | 3 | ... | 31

Cada celda de día trae las marcas del huellero separadas por saltos
de línea. El huellero suele repetir la misma marca varias veces:

    "06:07\\n06:07\\n13:15\\n13:15"   -> entrada 06:07, salida 13:15
    "06:08\\n13:19"                  -> entrada 06:08, salida 13:19
    "13:19"                          -> una sola marca: incompleta
    "06:07\\n06:07"                  -> misma hora repetida: incompleta
    ""                               -> sin registro (no se guarda)

REGLA: se toma la PRIMERA y la ÚLTIMA marca distintas del día.
Si solo hay una hora distinta, el día queda marcado como
`incompleta`, porque falta la entrada o la salida y no se puede
calcular la jornada.
"""
import calendar
import re
import unicodedata
from datetime import date, datetime, time
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HOJA_DATOS = "formato"
ALIAS_HOJA = {"formato", "formatolleno", "asistencia", "datos", "hoja1"}

MESES = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
         "agosto", "septiembre", "octubre", "noviembre", "diciembre"]

VERDE = "16412B"
VERDE2 = "2F7D4F"
CREMA = "EAE5D9"

# hh:mm  ·  hh:mm:ss  ·  h:mm
_RE_HORA = re.compile(r"(\d{1,2}):(\d{2})(?::(\d{2}))?")


def dias_del_mes(anio: int, mes: int) -> int:
    """Días que tuvo ese mes. Contempla los años bisiestos."""
    return calendar.monthrange(int(anio), int(mes))[1]


def nombre_mes(mes: int) -> str:
    return MESES[int(mes) - 1] if 1 <= int(mes) <= 12 else str(mes)


def _norm(texto) -> str:
    if texto is None:
        return ""
    t = unicodedata.normalize("NFKD", str(texto).strip().lower())
    t = "".join(c for c in t if not unicodedata.combining(c))
    for ch in (" ", "_", ".", "-", "/"):
        t = t.replace(ch, "")
    return t


def _txt(v):
    if v is None:
        return None
    t = str(v).strip()
    return t or None


def extraer_marcas(celda) -> list[time]:
    """
    Saca todas las horas de una celda, en el orden en que aparecen.
    Tolera saltos de línea, comas, espacios y celdas con formato de hora.
    """
    if celda is None:
        return []

    # Celda con formato de hora de Excel
    if isinstance(celda, time):
        return [celda]
    if isinstance(celda, datetime):
        return [celda.time()]

    texto = str(celda).strip()
    if not texto:
        return []

    marcas = []
    for h, m, s in _RE_HORA.findall(texto):
        try:
            hh, mm = int(h), int(m)
            ss = int(s) if s else 0
            if 0 <= hh <= 23 and 0 <= mm <= 59 and 0 <= ss <= 59:
                marcas.append(time(hh, mm, ss))
        except ValueError:
            continue
    return marcas


# Por debajo de estos minutos, la jornada es casi seguro un error de
# marcación (marcó varias veces al entrar y ninguna al salir).
MINUTOS_SOSPECHOSOS = 30


def jornada(marcas: list[time]) -> dict:
    """
    Convierte las marcas de un día en entrada, salida y duración.

    Toma la primera y la última hora DISTINTAS. Si todas las marcas
    son la misma hora, se trata como una sola marcación: falta la
    entrada o la salida, así que no hay jornada que calcular.

    Una jornada de pocos minutos también se marca como incompleta:
    suele ser alguien que marcó dos veces seguidas al entrar.
    """
    if not marcas:
        return {"estado": "sin_registro", "entrada": None, "salida": None,
                "minutos": None, "n_marcas": 0}

    ordenadas = sorted(marcas)
    distintas = sorted({(t.hour, t.minute, t.second) for t in ordenadas})

    if len(distintas) < 2:
        # Una sola hora (aunque el huellero la haya repetido)
        return {"estado": "incompleta", "entrada": ordenadas[0],
                "salida": None, "minutos": None, "n_marcas": len(marcas)}

    entrada, salida = ordenadas[0], ordenadas[-1]
    minutos = ((salida.hour * 60 + salida.minute) -
               (entrada.hour * 60 + entrada.minute))

    if minutos < MINUTOS_SOSPECHOSOS:
        # Marcas muy juntas: no es una jornada real
        return {"estado": "incompleta", "entrada": entrada, "salida": salida,
                "minutos": minutos, "n_marcas": len(marcas)}

    return {"estado": "completo", "entrada": entrada, "salida": salida,
            "minutos": minutos, "n_marcas": len(marcas)}


def _buscar_hoja(wb):
    for nombre in wb.sheetnames:
        if _norm(nombre) in ALIAS_HOJA:
            return wb[nombre]
    return wb[wb.sheetnames[0]]


def leer_excel(contenido: bytes, anio: int, mes: int) -> tuple[dict, list[str]]:
    """
    Lee el reporte del huellero.

    Devuelve (resultado, advertencias) donde resultado es:
      {"trabajadores": [{codigo, nombre, dias: [ {dia, ...} ]}],
       "dias_mes": 31, "resumen": {...}}
    """
    advertencias: list[str] = []
    total_dias = dias_del_mes(anio, mes)

    try:
        wb = load_workbook(BytesIO(contenido), data_only=True)
    except Exception as e:
        return {}, [f"No se pudo abrir el archivo: {e}"]

    ws = _buscar_hoja(wb)
    filas = list(ws.iter_rows(values_only=True))
    if len(filas) < 2:
        return {}, ["El archivo no tiene filas de datos."]

    # --- Cabecera: ubicar código, nombre y las columnas de día ---
    cabecera = filas[0]
    col_codigo = col_nombre = None
    cols_dia: dict[int, int] = {}          # índice de columna -> día

    for i, valor in enumerate(cabecera):
        etiqueta = _norm(valor)
        if etiqueta in ("employeeid", "id", "codigo", "cedula", "documento"):
            col_codigo = i
        elif etiqueta in ("name", "nombre", "empleado", "trabajador"):
            col_nombre = i
        else:
            texto = _txt(valor)
            if texto and texto.isdigit():
                d = int(texto)
                if 1 <= d <= 31:
                    cols_dia[i] = d

    if col_codigo is None and col_nombre is None:
        return {}, ["No se encontraron las columnas «Employee ID» ni «Name». "
                    "Descarga el formato desde el módulo y úsalo como base."]
    if not cols_dia:
        return {}, ["No se encontraron columnas de días (1, 2, 3…) en la fila 1."]

    # Avisar si el archivo no coincide con el mes elegido
    dias_archivo = max(cols_dia.values())
    if dias_archivo > total_dias:
        advertencias.append(
            f"El archivo trae {dias_archivo} días pero {nombre_mes(mes)} de {anio} "
            f"tiene {total_dias}. Los días sobrantes se ignoran.")
    elif dias_archivo < total_dias:
        advertencias.append(
            f"El archivo trae {dias_archivo} días y {nombre_mes(mes)} de {anio} "
            f"tiene {total_dias}. Los días faltantes quedan sin registro.")

    # --- Filas de trabajadores ---
    trabajadores: list[dict] = []
    vistos: set[str] = set()
    n_completos = n_incompletas = n_sin = 0

    for nfila, fila in enumerate(filas[1:], start=2):
        if fila is None:
            continue

        codigo = _txt(fila[col_codigo]) if col_codigo is not None and col_codigo < len(fila) else None
        nombre = _txt(fila[col_nombre]) if col_nombre is not None and col_nombre < len(fila) else None

        if not codigo and not nombre:
            continue
        if not nombre:
            nombre = f"Sin nombre ({codigo})"
        if not codigo:
            codigo = nombre          # sin ID, el nombre hace de llave

        clave = codigo.lower()
        if clave in vistos:
            advertencias.append(
                f"Fila {nfila}: el código «{codigo}» está repetido. Se omite.")
            continue
        vistos.add(clave)

        dias: list[dict] = []
        for indice, numero_dia in cols_dia.items():
            if numero_dia > total_dias:
                continue
            celda = fila[indice] if indice < len(fila) else None
            marcas = extraer_marcas(celda)
            j = jornada(marcas)

            if j["estado"] == "sin_registro":
                n_sin += 1
                continue        # los días sin marcas no se guardan

            if j["estado"] == "completo":
                n_completos += 1
            else:
                n_incompletas += 1

            dias.append({
                "dia": numero_dia,
                "fecha": date(int(anio), int(mes), numero_dia),
                "entrada": j["entrada"],
                "salida": j["salida"],
                "minutos": j["minutos"],
                "estado": j["estado"],
                "n_marcas": j["n_marcas"],
                "marcas": [t.strftime("%H:%M:%S") for t in marcas],
            })

        trabajadores.append({"codigo": codigo, "nombre": nombre,
                             "fila_excel": nfila, "dias": dias})

    if not trabajadores:
        return {}, ["No se encontró ningún trabajador en el archivo."]

    if n_incompletas:
        advertencias.append(
            f"{n_incompletas} día(s) sin jornada calculable: una sola marcación, "
            f"o entrada y salida a menos de {MINUTOS_SOSPECHOSOS} minutos. "
            f"Quedan registrados y se listan en el análisis para revisarlos.")

    return {
        "trabajadores": trabajadores,
        "dias_mes": total_dias,
        "resumen": {"trabajadores": len(trabajadores),
                    "dias_completos": n_completos,
                    "dias_incompletos": n_incompletas,
                    "dias_sin_registro": n_sin},
    }, advertencias


# ============================================================
#  Generación del formato
# ============================================================

def generar_formato(anio: int, mes: int,
                    trabajadores: list[dict] | None = None) -> bytes:
    """
    Genera el Excel con una columna por día del mes elegido.
    Febrero de un año bisiesto trae 29; los meses de 30 traen 30.

    Si se pasan trabajadores (por ejemplo los del mes anterior),
    se precargan su código y su nombre.
    """
    total = dias_del_mes(anio, mes)
    personas = trabajadores or []

    wb = Workbook()

    # ---------- Hoja de datos ----------
    ws = wb.active
    ws.title = HOJA_DATOS
    ws.append(["Employee ID", "Name"] + [str(d) for d in range(1, total + 1)])
    for p in personas:
        ws.append([p.get("codigo"), p.get("nombre")])

    relleno = PatternFill("solid", fgColor=VERDE)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF", size=10)
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 34
    for d in range(total):
        ws.column_dimensions[get_column_letter(3 + d)].width = 11

    # Las celdas de día llevan varias horas: se ven mejor con ajuste
    for fila in range(2, max(2, len(personas) + 1) + 1):
        for col in range(3, total + 3):
            ws.cell(row=fila, column=col).alignment = Alignment(
                wrap_text=True, vertical="top", horizontal="center")

    ws.freeze_panes = "C2"

    # ---------- Instrucciones ----------
    guia = wb.create_sheet("instrucciones")
    lineas = [
        (f"PalmaData · Asistencia · {nombre_mes(mes).capitalize()} {anio}", "titulo"),
        ("", ""),
        (f"Este formato tiene {total} columnas de día, "
         f"las que tuvo {nombre_mes(mes)} de {anio}.", ""),
        ("", ""),
        ("Cómo llenarlo", "sub"),
        ("1. Exporta el reporte del huellero para este mes.", ""),
        ("2. Copia los datos y pégalos en la hoja «formato», desde la celda A2.", ""),
        ("   Usa Pegado especial → Valores.", ""),
        ("3. Sube el archivo desde Asistencia → Cargar datos, eligiendo la", ""),
        ("   empresa, el año y el mes.", ""),
        ("", ""),
        ("Estructura", "sub"),
        ("· Columna A: Employee ID (el código del huellero).", ""),
        ("· Columna B: Name (el nombre del trabajador).", ""),
        ("· De la columna C en adelante: un día del mes cada una.", ""),
        ("", ""),
        ("Las marcaciones", "sub"),
        ("Cada celda de día lleva las marcas de esa persona ese día,", ""),
        ("una debajo de otra. El huellero suele repetir la misma marca:", ""),
        ("", ""),
        ("   06:07                     entrada 06:07", ""),
        ("   06:07        se lee →     salida  13:15", ""),
        ("   13:15                     jornada 7h 08m", ""),
        ("   13:15", ""),
        ("", ""),
        ("El sistema toma la primera y la última hora distintas del día.", ""),
        ("No importa cuántas marcas haya: 2, 4 o 10, el resultado es el mismo.", ""),
        ("", ""),
        ("Si un día solo tiene una hora (o la misma repetida), queda marcado", ""),
        ("como incompleto: falta la entrada o la salida, así que no se puede", ""),
        ("calcular la jornada. Lo mismo si entrada y salida quedan a menos de", ""),
        ("30 minutos, que suele ser alguien que marcó dos veces al entrar.", ""),
        ("El registro se guarda igual y aparece en el análisis para revisarlo.", ""),
        ("", ""),
        ("Las celdas vacías son días sin marcación (domingos, festivos,", ""),
        ("ausencias) y no entran en los promedios.", ""),
        ("", ""),
        ("Recargar el mismo mes", "sub"),
        ("Si vuelves a subir la misma empresa, año y mes, los datos anteriores", ""),
        ("se reemplazan. Sirve para corregir un archivo equivocado.", ""),
    ]
    for texto, tipo in lineas:
        guia.append([texto])
        celda = guia.cell(row=guia.max_row, column=1)
        if tipo == "titulo":
            celda.font = Font(bold=True, size=14, color=VERDE)
        elif tipo == "sub":
            celda.font = Font(bold=True, size=11, color=VERDE2)
    guia.column_dimensions["A"].width = 78

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
