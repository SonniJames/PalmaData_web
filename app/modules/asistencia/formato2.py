"""
PalmaData · Asistencia · Formato 2
==================================
El huellero antiguo exporta un archivo distinto: en vez de una matriz
de días, trae UNA FILA POR TRABAJADOR Y FECHA, con la entrada y la
salida ya separadas en columnas.

    ID | Nombre | Departamento | Fecha | Entrada | Salida
    31 | Nicolasvasques | Empresa | 2026-08-05 | 05:50 | 13:06

Casos que trae:
    entrada y salida  -> jornada calculada
    solo entrada      -> a revisar (falta la salida)
    solo salida       -> a revisar (falta la entrada)
    ninguna           -> sin registro: el día no se guarda, pero la
                         persona sí queda en el padrón del huellero

El resultado sale con la MISMA estructura que el formato 1, así que
todo lo demás del módulo (base de datos, análisis, pantallas) es
idéntico para los dos formatos.
"""
import re
from datetime import date, datetime, time
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .excel_loader import (CREMA, MINUTOS_SOSPECHOSOS, VERDE, VERDE2,
                           _norm, _txt, dias_del_mes, nombre_mes)

HOJA_DATOS = "formato"
ALIAS_HOJA = {"formato", "formatolleno", "asistencia", "datos", "hoja1"}

COLUMNAS = ["ID", "Nombre", "Departamento", "Fecha", "Entrada", "Salida"]

# encabezado normalizado -> campo interno
MAPA_COLUMNAS = {
    "id": "codigo", "employeeid": "codigo", "codigo": "codigo",
    "nombre": "nombre", "name": "nombre", "empleado": "nombre",
    "departamento": "departamento", "department": "departamento",
    "area": "departamento",
    "fecha": "fecha", "date": "fecha", "dia": "fecha",
    "entrada": "entrada", "horaentrada": "entrada", "checkin": "entrada",
    "salida": "salida", "horasalida": "salida", "checkout": "salida",
}

_RE_HORA = re.compile(r"^(\d{1,2}):(\d{2})(?::(\d{2}))?$")
_RE_FECHA = re.compile(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})")
_RE_FECHA_INV = re.compile(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})")


def _hora(valor) -> time | None:
    """Convierte una celda a hora. Acepta texto 'HH:MM' o formato de Excel."""
    if valor is None:
        return None
    if isinstance(valor, time):
        return valor
    if isinstance(valor, datetime):
        return valor.time()
    texto = str(valor).strip()
    if not texto or texto.lower() in ("none", "-", "—", "null"):
        return None
    m = _RE_HORA.match(texto)
    if not m:
        return None
    try:
        hh, mm = int(m.group(1)), int(m.group(2))
        ss = int(m.group(3)) if m.group(3) else 0
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return time(hh, mm, ss)
    except ValueError:
        pass
    return None


def _fecha(valor) -> date | None:
    """Convierte una celda a fecha. Acepta 'AAAA-MM-DD', 'DD/MM/AAAA' o fecha de Excel."""
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = str(valor).strip()
    if not texto:
        return None
    m = _RE_FECHA.match(texto)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    m = _RE_FECHA_INV.match(texto)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    return None


def _buscar_hoja(wb):
    for nombre in wb.sheetnames:
        if _norm(nombre) in ALIAS_HOJA:
            return wb[nombre]
    return wb[wb.sheetnames[0]]


def jornada_desde_columnas(entrada: time | None, salida: time | None) -> dict:
    """
    Arma la jornada con la entrada y la salida ya separadas.

    Aquí no hay que deducir nada: el huellero ya las distingue.
    Solo falta validar que estén las dos y que la duración tenga sentido.
    """
    if entrada is None and salida is None:
        return {"estado": "sin_registro", "entrada": None, "salida": None,
                "minutos": None, "n_marcas": 0}

    if entrada is None or salida is None:
        return {"estado": "incompleta", "entrada": entrada, "salida": salida,
                "minutos": None, "n_marcas": 1}

    minutos = ((salida.hour * 60 + salida.minute) -
               (entrada.hour * 60 + entrada.minute))

    if minutos < 0:
        # Salida antes que la entrada: turno nocturno o dato invertido
        return {"estado": "incompleta", "entrada": entrada, "salida": salida,
                "minutos": minutos, "n_marcas": 2}

    if minutos < MINUTOS_SOSPECHOSOS:
        return {"estado": "incompleta", "entrada": entrada, "salida": salida,
                "minutos": minutos, "n_marcas": 2}

    return {"estado": "completo", "entrada": entrada, "salida": salida,
            "minutos": minutos, "n_marcas": 2}


def leer_excel(contenido: bytes, anio: int, mes: int) -> tuple[dict, list[str]]:
    """
    Lee el archivo del formato 2.

    Devuelve la MISMA estructura que el formato 1, para que el resto
    del módulo no tenga que distinguirlos.
    """
    advertencias: list[str] = []
    total_dias = dias_del_mes(anio, mes)

    try:
        wb = load_workbook(BytesIO(contenido), data_only=True)
    except Exception as e:
        return {}, [f"No se pudo abrir el archivo: {e}"]

    ws = _buscar_hoja(wb)
    filas = list(ws.iter_rows(values_only=True))
    if len(filas) < 2:
        return {}, ["El archivo no tiene filas de datos."]

    # --- Cabecera ---
    cabecera = filas[0]
    mapa: dict[int, str] = {}
    for i, valor in enumerate(cabecera):
        campo = MAPA_COLUMNAS.get(_norm(valor))
        if campo:
            mapa[i] = campo

    faltantes = [c for c in ("codigo", "fecha") if c not in mapa.values()]
    if faltantes:
        nombres = {"codigo": "ID", "fecha": "Fecha"}
        return {}, [f"Faltan columnas obligatorias: "
                    f"{', '.join(nombres[c] for c in faltantes)}. "
                    f"Descarga el formato 2 desde el módulo y úsalo como base."]

    inverso = {v: k for k, v in mapa.items()}

    # --- Filas ---
    personas: dict[str, dict] = {}
    n_completos = n_incompletas = n_sin = 0
    fuera_periodo = 0
    sin_fecha = 0
    duplicados = 0

    def celda(fila, campo):
        i = inverso.get(campo)
        return fila[i] if (i is not None and i < len(fila)) else None

    for nfila, fila in enumerate(filas[1:], start=2):
        if fila is None or all(v in (None, "") for v in fila):
            continue

        codigo = _txt(celda(fila, "codigo"))
        nombre = _txt(celda(fila, "nombre"))
        if not codigo and not nombre:
            continue
        if not nombre:
            nombre = f"Sin nombre ({codigo})"
        if not codigo:
            codigo = nombre

        f = _fecha(celda(fila, "fecha"))
        if f is None:
            sin_fecha += 1
            continue

        if f.year != int(anio) or f.month != int(mes):
            fuera_periodo += 1
            continue

        entrada = _hora(celda(fila, "entrada"))
        salida = _hora(celda(fila, "salida"))
        j = jornada_desde_columnas(entrada, salida)

        # La persona se registra SIEMPRE, aunque ese día no haya marcado.
        # Así queda en el padrón del huellero y aparece en el análisis,
        # igual que en el formato 1. Solo el DÍA sin marcas se omite.
        p = personas.setdefault(codigo, {
            "codigo": codigo, "nombre": nombre,
            "fila_excel": nfila, "dias": {},
        })
        p["nombre"] = nombre        # el último nombre visto manda

        if j["estado"] == "sin_registro":
            n_sin += 1
            continue

        if f.day in p["dias"]:
            duplicados += 1
            continue

        marcas = [t.strftime("%H:%M:%S") for t in (entrada, salida) if t]
        p["dias"][f.day] = {
            "dia": f.day, "fecha": f,
            "entrada": j["entrada"], "salida": j["salida"],
            "minutos": j["minutos"], "estado": j["estado"],
            "n_marcas": j["n_marcas"], "marcas": marcas,
            "departamento": _txt(celda(fila, "departamento")),
        }

        if j["estado"] == "completo":
            n_completos += 1
        else:
            n_incompletas += 1

    if not personas:
        return {}, ["No se encontró ningún registro válido para "
                    f"{nombre_mes(mes)} de {anio}. Revisa que las fechas "
                    "del archivo correspondan al período seleccionado."]

    # --- Avisos ---
    if fuera_periodo:
        advertencias.append(
            f"{fuera_periodo} fila(s) con fecha fuera de {nombre_mes(mes)} "
            f"de {anio}. Se omitieron: revisa que el archivo sea del período correcto.")
    if sin_fecha:
        advertencias.append(f"{sin_fecha} fila(s) sin fecha legible. Se omitieron.")
    if duplicados:
        advertencias.append(
            f"{duplicados} fila(s) repetidas (mismo trabajador y fecha). "
            f"Se conservó la primera de cada una.")
    if n_incompletas:
        advertencias.append(
            f"{n_incompletas} día(s) sin jornada calculable: falta la entrada "
            f"o la salida, o la duración es menor a {MINUTOS_SOSPECHOSOS} minutos. "
            f"Quedan registrados y se listan en el análisis.")

    trabajadores = [{"codigo": p["codigo"], "nombre": p["nombre"],
                     "fila_excel": p["fila_excel"],
                     "dias": sorted(p["dias"].values(), key=lambda d: d["dia"])}
                    for p in personas.values()]

    return {
        "trabajadores": trabajadores,
        "dias_mes": total_dias,
        "resumen": {"trabajadores": len(trabajadores),
                    "dias_completos": n_completos,
                    "dias_incompletos": n_incompletas,
                    "dias_sin_registro": n_sin},
    }, advertencias


# ============================================================
#  Generación del formato en blanco
# ============================================================

def generar_formato(anio: int, mes: int,
                    trabajadores: list[dict] | None = None) -> bytes:
    """
    Formato 2 en blanco: una fila por trabajador y fecha.

    Si se pasan trabajadores, se precarga una fila por cada persona
    y cada día del mes, para que solo haya que llenar las horas.
    """
    personas = trabajadores or []
    total = dias_del_mes(anio, mes)

    wb = Workbook()
    ws = wb.active
    ws.title = HOJA_DATOS
    ws.append(COLUMNAS)

    for p in personas:
        for d in range(1, total + 1):
            ws.append([p.get("codigo"), p.get("nombre"), None,
                       f"{anio}-{int(mes):02d}-{d:02d}", None, None])

    relleno = PatternFill("solid", fgColor=VERDE)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF", size=10)
        celda.fill = relleno
        celda.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    for letra, ancho in (("A", 12), ("B", 32), ("C", 18),
                         ("D", 14), ("E", 12), ("F", 12)):
        ws.column_dimensions[letra].width = ancho
    ws.freeze_panes = "A2"

    # ---------- Instrucciones ----------
    guia = wb.create_sheet("instrucciones")
    lineas = [
        (f"PalmaData · Asistencia · Formato 2 · "
         f"{nombre_mes(mes).capitalize()} {anio}", "titulo"),
        ("", ""),
        ("Este formato es el del huellero antiguo: una fila por", ""),
        ("trabajador y fecha, con la entrada y la salida separadas.", ""),
        ("", ""),
        ("Columnas", "sub"),
        ("· ID: el código del trabajador en el huellero. Obligatorio.", ""),
        ("· Nombre: nombre de la persona.", ""),
        ("· Departamento: opcional, se guarda tal cual.", ""),
        ("· Fecha: AAAA-MM-DD (por ejemplo 2026-08-05). Obligatorio.", ""),
        ("· Entrada: hora de entrada en formato HH:MM.", ""),
        ("· Salida: hora de salida en formato HH:MM.", ""),
        ("", ""),
        ("Cómo se leen los datos", "sub"),
        ("   Entrada 05:50 y salida 13:06  ->  jornada 7h 16m", ""),
        ("   Solo entrada                  ->  a revisar", ""),
        ("   Solo salida                   ->  a revisar", ""),
        ("   Ninguna de las dos            ->  sin registro", ""),
        ("", ""),
        (f"Solo se cargan las filas cuya fecha sea de "
         f"{nombre_mes(mes)} de {anio}.", ""),
        ("Las de otros meses se omiten y el sistema te avisa cuántas fueron.", ""),
        ("", ""),
        ("Si un trabajador aparece dos veces el mismo día, se conserva", ""),
        ("la primera fila y se avisa.", ""),
        ("", ""),
        ("Recargar el mismo período", "sub"),
        ("Si vuelves a subir la misma empresa, zona, año y mes, los datos", ""),
        ("anteriores se reemplazan. Las otras zonas no se tocan.", ""),
    ]
    for texto, tipo in lineas:
        guia.append([texto])
        celda = guia.cell(row=guia.max_row, column=1)
        if tipo == "titulo":
            celda.font = Font(bold=True, size=14, color=VERDE)
        elif tipo == "sub":
            celda.font = Font(bold=True, size=11, color=VERDE2)
    guia.column_dimensions["A"].width = 74

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
